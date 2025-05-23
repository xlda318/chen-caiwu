# 导入需要的库
# pandas 用于数据处理和分析
import pandas as pd
# openpyxl 用于操作 Excel 文件（读取/写入）
from openpyxl import load_workbook
# 从 openpyxl.utils 模块导入获取列字母的函数（如 1->A, 2->B）
from openpyxl.utils import get_column_letter


# 定义一个函数：将表头和数据写入 Excel 工作表
def write_data(worksheet, headers, data):
    """
    参数说明：
    worksheet: 要写入数据的工作表对象（openpyxl 的 Worksheet 对象）
    headers: 表头列表（例如 ['姓名', '年龄', '成绩']）
    data: 要写入的数据（pandas 的 DataFrame 格式）
    """
    # 写入表头：从第 1 行开始，依次写入每个表头字段
    for col_idx, header in enumerate(headers, 1):  # enumerate 生成索引和值，从 1 开始计数
        worksheet.cell(row=1, column=col_idx, value=header)  # cell 方法指定单元格位置并赋值

    # 写入数据：从第 2 行开始（第 1 行已写入表头）
    for row_idx, row_data in enumerate(data.reset_index().values, 2):
        # reset_index() 将数据透视表的索引（如机床号、图号）转换为普通列
        # values 属性获取数据的数值部分（二维数组）
        # enumerate 从 2 开始计数，因为第 1 行是表头
        for col_idx, value in enumerate(row_data, 1):  # 遍历每行的每个数据值
            worksheet.cell(row=row_idx, column=col_idx, value=value)  # 写入单元格


# 主函数：创建数据透视表并保存到原 Excel 文件
def create_pivot_table():
    # 配置文件路径和工作表名称
    file_path = r"D:\code\家里电脑\工时\1月车工提交.xlsx"  # r 表示原始字符串，避免 \ 转义问题
    sheet_name = "源文件 (整理)"  # 数据源所在的工作表名称
    output_sheet = "工时透视表"  # 透视结果保存的工作表名称

    # ---------------------- 1. 读取数据 ----------------------
    try:
        # 使用 pandas 读取 Excel 文件中的指定工作表
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        print(f"数据已成功加载，共有{len(df)}行，{len(df.columns)}列")  # 打印数据行数和列数
    except Exception as e:  # 捕获读取文件时的错误
        print(f"读取文件时出错: {e}")  # 打印错误信息
        return  # 出错后直接退出函数

    # ---------------------- 2. 查看列名（调试用）----------------------
    print("\n数据所有列名:")  # 换行后打印列名列表
    for col in df.columns:  # 遍历所有列名
        print(f"- {col}")  # 打印每个列名（检查是否存在空格、特殊字符等）

    # ---------------------- 3. 处理列名（关键步骤）----------------------
    # 原始列名可能包含空格（如 '图       号'），需要映射为规范名称
    column_mapping = {
        '图       号': '图号',   # 原列名有多个空格，映射为 '图号'
        '零件名称': '零件名称',   # 名称正确，直接映射（可省略，但为了清晰保留）
        '机床号': '机床号',       # 名称正确，直接映射
        '完成数量': '完成数量',     # 名称正确，直接映射
        '合计': '合计'           # 名称正确，直接映射
    }

    # 获取数据中实际存在的列名（过滤掉不存在的键）
    existing_columns = [col for col in df.columns if col in column_mapping.keys()]
    # 根据实际存在的列名创建有效的映射关系
    valid_columns = {col: column_mapping[col] for col in existing_columns}

    # 检查是否所有必要列都存在（避免后续操作出错）
    required_columns = list(column_mapping.values())  # 期望的列名列表
    missing_columns = [col for col in required_columns if col not in valid_columns.values()]
    if missing_columns:  # 如果有缺失的列
        print(f"错误: 以下必要列不存在: {', '.join(missing_columns)}")
        return  # 出错后退出函数

    # 重命名列（将原始列名替换为规范名称）
    df = df.rename(columns=valid_columns)  # rename 方法返回新的 DataFrame，原数据不变

    # ---------------------- 4. 数据筛选与类型转换 ----------------------
    # 选择需要的列（使用规范后的列名）
    selected_columns = list(valid_columns.values())  # 获取规范后的列名列表
    df_selected = df[selected_columns].copy()  # 复制数据，避免修改原始数据

    # 将字符串类型的数值列转换为数字类型（如 '完成数量' 可能是文本格式）
    df_selected['完成数量'] = pd.to_numeric(df_selected['完成数量'], errors='coerce')
    df_selected['合计'] = pd.to_numeric(df_selected['合计'], errors='coerce')
    # 处理转换后可能出现的缺失值（NaN），用 0 填充
    df_selected['完成数量'] = df_selected['完成数量'].fillna(0)
    df_selected['合计'] = df_selected['合计'].fillna(0)

    # ---------------------- 5. 创建数据透视表 ----------------------
    pivot_table = pd.pivot_table(
        df_selected,                # 要透视的数据源
        index=['机床号', '图号', '零件名称'],  # 行标签（按这三列分组）
        values=['完成数量', '合计'],       # 要汇总的数值列
        aggfunc={'完成数量': 'sum', '合计': 'sum'}  # 聚合函数：求和
    )
    # 重命名透视表的列（更直观）
    pivot_table = pivot_table.rename(columns={'完成数量': '数量', '合计': '工时'})

    # 计算工时占比（某行工时 / 总工时 * 100%）
    pivot_table['占比'] = pivot_table['工时'] / pivot_table['工时'].sum() * 100
    # 将占比格式化为百分比字符串（保留两位小数）
    pivot_table['占比'] = pivot_table['占比'].map('{:.2f}%'.format)

    # ---------------------- 6. 保存结果到原 Excel 文件 ----------------------
    try:
        # 加载原 Excel 文件（保留原有内容）
        wb = load_workbook(file_path)
        # 如果结果工作表已存在，先删除（避免重复）
        if output_sheet in wb.sheetnames:
            del wb[output_sheet]
        # 创建新工作表
        ws = wb.create_sheet(output_sheet)

        # 定义表头（与透视表列顺序一致）
        headers = ['机床号', '图号', '零件名称', '数量', '工时', '占比']
        # 调用函数写入表头和数据
        write_data(ws, headers, pivot_table)

        # 自动调整列宽（让内容完整显示）
        for col_idx in range(1, len(headers) + 1):  # 遍历所有列（从第 1 列到最后一列）
            max_length = 0  # 记录当前列的最大内容长度
            column = get_column_letter(col_idx)  # 获取列字母（如 1->A, 2->B）
            # 遍历当前列的所有行（包括表头和数据）
            for row in range(1, ws.max_row + 1):
                # 获取单元格内容并转换为字符串，计算长度
                cell_value = str(ws[f"{column}{row}"].value) if ws[f"{column}{row}"].value else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            # 设置列宽（内容长度 + 2 作为缓冲）
            ws.column_dimensions[column].width = max_length + 2

        # 保存修改后的 Excel 文件
        wb.save(file_path)
        print(f"\n数据透视表已保存至: {file_path} 的 '{output_sheet}' 工作表")
    except PermissionError as pe:  # 捕获文件被占用的错误
        print(f"保存文件时权限错误，可能文件被其他程序占用，请关闭相关程序后重试。错误详情: {pe}")
    except Exception as e:  # 捕获其他保存错误
        print(f"保存透视表时出错: {e}")


# 当直接运行脚本时，执行主函数
if __name__ == "__main__":
    create_pivot_table()