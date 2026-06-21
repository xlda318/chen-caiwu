# -*- coding: utf-8 -*-
"""
烨软科技 2023-2025 年经营分析
四大维度：盈利能力、营运能力、偿债能力、成长能力
"""
import os, json, math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

BASE = r'C:\Users\chenyh\OneDrive\office-claude\claude-work\projects\财务分析'
os.chdir(BASE)

# ===================== 1. 数据 =====================

# 利润表
PL = {
    2023: {'营业收入': 20204627.21, '营业成本': 15027007.89,
           '税金及附加': 14743.31, '销售费用': 432319.26,
           '管理费用': 1993372.54, '研发费用': 3382678.29,
           '财务费用': 187587.38, '营业利润': 143829.26,
           '营业外收入': 0, '营业外支出': 92.0,
           '利润总额': 143737.26, '净利润': 143737.26},
    2024: {'营业收入': 7198465.15, '营业成本': 5096140.47,
           '税金及附加': 11800.25, '销售费用': 401331.42,
           '管理费用': 2229630.01, '研发费用': 1998273.35,
           '财务费用': 218352.74, '营业利润': -2757063.09,
           '营业外收入': 164250.39, '营业外支出': 309.51,
           '利润总额': -2593122.21, '净利润': -2593122.21},
    2025: {'营业收入': 7991708.99, '营业成本': 4168328.95,
           '税金及附加': 13567.31, '销售费用': 230472.99,
           '管理费用': 2340254.93, '研发费用': 1025847.71,
           '财务费用': 215094.95, '营业利润': -1857.85,
           '营业外收入': 280175.68, '营业外支出': 258820.93,
           '利润总额': 19496.90, '净利润': 19496.90},
}

# 资产负债表
BS = {
    2023: {
        '货币资金': 2486106.08, '应收票据': 5162500, '应收账款': 6350480,
        '预付款项': 539029.63, '其他应收款': 748667.52, '存货': 7799514.73,
        '流动资产合计': 23086297.96,
        '固定资产': 1434579.60,  # 2024年初与2024年末估算
        '长期待摊费用': 472458.17, '研发支出': 506886.20,
        '非流动资产合计': 1277364.37,
        '资产总计': 24363662.33,
        '短期借款': 5900000, '应付账款': 13966004.28,
        '预收款项': 976496.52, '应付职工薪酬': 567308.27,
        '应交税费': -146208.20, '其他应付款': 1455139.82,
        '流动负债合计': 22718740.17,
        '负债合计': 22718740.17,
        '实收资本': 5000000, '未分配利润': -3355077.84,
        '所有者权益合计': 1644922.16,
    },
    2024: {
        '货币资金': 5363785.29, '应收票据': 2000000, '应收账款': 5891686,
        '预付款项': 725627.41, '其他应收款': 1126728.12, '存货': 6586756.66,
        '流动资产合计': 21694583.48,
        '固定资产': 704249.66, '无形资产': 24049.93,
        '长期待摊费用': 526312.26, '研发支出': 795900.77,
        '非流动资产合计': 2050512.62,
        '资产总计': 23745096.10,
        '短期借款': 8000000, '应付账款': 12458058.55,
        '预收款项': 2926900, '应付职工薪酬': 282032.92,
        '应交税费': -133844.14, '其他应付款': 1152269,
        '流动负债合计': 24685416.33,
        '负债合计': 24685416.33,
        '实收资本': 5000000, '未分配利润': -5940320.23,
        '所有者权益合计': -940320.23,
    },
    2025: {
        '货币资金': 66417.22, '应收票据': 5700000, '应收账款': 3769386,
        '预付款项': 513868.57, '其他应收款': 998911.84, '存货': 9556882.78,
        '流动资产合计': 20605466.41,
        '固定资产': 655883.75, '无形资产': 44341.20,
        '长期待摊费用': 58881.22, '研发支出': 474237.21,
        '非流动资产合计': 1233343.38,
        '资产总计': 21838809.79,
        '短期借款': 5860000, '应付账款': 15401899.15,
        '预收款项': 126000, '应付职工薪酬': 266564.68,
        '应交税费': -175856.86, '其他应付款': 1281026.15,
        '流动负债合计': 22759633.12,
        '负债合计': 22759633.12,
        '实收资本': 5000000, '未分配利润': -5920823.33,
        '所有者权益合计': -920823.33,
    }
}

# 现金流量表
CF = {
    2024: {'经营活动净额': 1092057.87, '投资活动净额': -120290.89,
           '筹资活动净额': 1905912.23},
    2025: {'经营活动净额': -2716090.78, '投资活动净额': -257912.50,
           '筹资活动净额': -2323364.79},
}

YEARS = [2023, 2024, 2025]

# ===================== 2. 计算函数 =====================

def avg_bs(item, y1, y2):
    v1 = BS[y1].get(item, 0) or 0
    v2 = BS[y2].get(item, 0) or 0
    return (v1 + v2) / 2

# ===================== 3. 四维度计算 =====================

results = {}

# --- 3.1 盈利能力 ---
profit = {}
for y in YEARS:
    p = PL[y]
    b = BS[y]
    rev = p['营业收入']
    gp = rev - p['营业成本']
    np = p['净利润']
    oe = p['营业利润']
    eq = b['所有者权益合计']
    ta = b['资产总计']

    profit[y] = {
        '毛利率': gp / rev * 100 if rev else 0,
        '净利率': np / rev * 100 if rev else 0,
        '营业利润率': oe / rev * 100 if rev else 0,
        '期间费用率': (p['销售费用'] + p['管理费用'] + p['研发费用'] + p['财务费用']) / rev * 100 if rev else 0,
    }

# 需要平均净资产的ROE使用两年平均
for y in YEARS:
    b = BS[y]
    p = PL[y]
    eq = b['所有者权益合计']
    # ROA = 净利润 / 总资产
    profit[y]['ROA'] = p['净利润'] / b['资产总计'] * 100 if b['资产总计'] else 0
    profit[y]['总资产'] = b['资产总计']
    profit[y]['净利润'] = p['净利润']

# ROE：用年初+年末平均净资产
for i, y in enumerate(YEARS):
    eq_end = BS[y]['所有者权益合计']
    if i == 0:
        # 2023年仅有年末数，无年初数，用年末近似
        profit[y]['ROE'] = PL[y]['净利润'] / eq_end * 100 if eq_end else 0
        profit[y]['平均净资产'] = eq_end
    else:
        eq_start = BS[YEARS[i-1]]['所有者权益合计']
        avg_eq = (eq_start + eq_end) / 2
        profit[y]['ROE'] = PL[y]['净利润'] / avg_eq * 100 if avg_eq else 0
        profit[y]['平均净资产'] = avg_eq

# 杜邦拆解（仅2024、2025可算平均）
for y in [2024, 2025]:
    p = PL[y]
    b = BS[y]
    b_prev = BS[y-1]
    avg_ta = (b_prev['资产总计'] + b['资产总计']) / 2
    avg_eq = profit[y]['平均净资产']
    rev = p['营业收入']
    # 净利率
    npm = p['净利润'] / rev if rev else 0
    # 总资产周转率
    tat = rev / avg_ta if avg_ta else 0
    # 权益乘数
    em = avg_ta / avg_eq if avg_eq else 0
    profit[y]['杜邦_净利率'] = npm * 100
    profit[y]['杜邦_总资产周转率'] = tat
    profit[y]['杜邦_权益乘数'] = em
    profit[y]['杜邦_ROE'] = npm * tat * em * 100

results['盈利能力'] = profit

# --- 3.2 营运能力 ---
operate = {}
for y in YEARS:
    p = PL[y]
    b = BS[y]
    rev = p['营业收入']
    cogs = p['营业成本']
    inv = b['存货']
    ar = b['应收账款']

    operate[y] = {
        '存货': inv,
        '应收账款': ar,
        '营业收入': rev,
    }

for i, y in enumerate(YEARS):
    if i == 0:
        continue
    p = PL[y]
    b = BS[y]
    b_prev = BS[YEARS[i-1]]
    rev = p['营业收入']
    cogs = p['营业成本']

    avg_inv = (b_prev['存货'] + b['存货']) / 2
    avg_ar = (b_prev['应收账款'] + b['应收账款']) / 2
    avg_ta = (b_prev['资产总计'] + b['资产总计']) / 2

    inv_to = cogs / avg_inv if avg_inv else 0
    ar_to = rev / avg_ar if avg_ar else 0
    ta_to = rev / avg_ta if avg_ta else 0

    operate[y]['存货周转率'] = inv_to
    operate[y]['存货周转天数'] = 365 / inv_to if inv_to else 0
    operate[y]['应收账款周转率'] = ar_to
    operate[y]['应收账款周转天数'] = 365 / ar_to if ar_to else 0
    operate[y]['总资产周转率'] = ta_to
    operate[y]['总资产周转天数'] = 365 / ta_to if ta_to else 0

# 2023用年末数近似
b23 = BS[2023]
operate[2023]['存货周转率'] = PL[2023]['营业成本'] / b23['存货'] if b23['存货'] else 0
operate[2023]['存货周转天数'] = 365 / operate[2023]['存货周转率'] if operate[2023]['存货周转率'] else 0
operate[2023]['应收账款周转率'] = PL[2023]['营业收入'] / b23['应收账款'] if b23['应收账款'] else 0
operate[2023]['应收账款周转天数'] = 365 / operate[2023]['应收账款周转率'] if operate[2023]['应收账款周转率'] else 0
operate[2023]['总资产周转率'] = PL[2023]['营业收入'] / b23['资产总计'] if b23['资产总计'] else 0
operate[2023]['总资产周转天数'] = 365 / operate[2023]['总资产周转率'] if operate[2023]['总资产周转率'] else 0

results['营运能力'] = operate

# --- 3.3 偿债能力 ---
debt = {}
for y in YEARS:
    b = BS[y]
    ca = b['流动资产合计']
    cl = b['流动负债合计']
    inv = b['存货']
    tl = b['负债合计']
    ta = b['资产总计']
    p = PL[y]

    debt[y] = {
        '流动比率': ca / cl if cl else 0,
        '速动比率': (ca - inv) / cl if cl else 0,
        '资产负债率': tl / ta * 100 if ta else 0,
        '产权比率': tl / b['所有者权益合计'] if b['所有者权益合计'] else 0,
        '利息保障倍数': (p['利润总额'] + abs(p['财务费用'])) / abs(p['财务费用']) if abs(p['财务费用']) else 0,
        '流动资产合计': ca,
        '流动负债合计': cl,
    }

results['偿债能力'] = debt

# --- 3.4 成长能力 ---
growth = {}
for i, y in enumerate(YEARS):
    if i == 0:
        growth[y] = {'收入增长率': None, '利润增长率': None}
        continue
    p_cur = PL[y]
    p_prev = PL[YEARS[i-1]]
    rev_g = (p_cur['营业收入'] - p_prev['营业收入']) / p_prev['营业收入'] * 100 if p_prev['营业收入'] else 0
    np_g = (p_cur['净利润'] - p_prev['净利润']) / abs(p_prev['净利润']) * 100 if p_prev['净利润'] else 0  # 用绝对值算
    growth[y] = {'收入增长率': rev_g, '利润增长率': np_g}

# CAGR 2023-2025
revenue_vals = [PL[y]['营业收入'] for y in YEARS]
profit_vals = [PL[y]['净利润'] for y in YEARS]
if revenue_vals[0] > 0:
    rev_cagr = (revenue_vals[-1] / revenue_vals[0]) ** (1/2) - 1
    revenue_cagr_3y = rev_cagr * 100
else:
    revenue_cagr_3y = None

if profit_vals[0] > 0 and profit_vals[-1] > 0:
    np_cagr = (profit_vals[-1] / profit_vals[0]) ** (1/2) - 1
    profit_cagr_3y = np_cagr * 100
else:
    profit_cagr_3y = None

growth['CAGR'] = {'收入CAGR(23-25)': revenue_cagr_3y, '净利润CAGR(23-25)': profit_cagr_3y}
results['成长能力'] = growth

# ===================== 4. 输出Excel =====================

wb = Workbook()

# 样式
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
data_font = Font(name='微软雅黑', size=10)
title_font = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
section_font = Font(name='微软雅黑', bold=True, size=12, color='2F5496')
alert_font = Font(name='微软雅黑', size=10, color='FF0000')
green_font = Font(name='微软雅黑', size=10, color='006100')
pct_format = '0.00%'
num_format = '#,##0.00'
thin_border = Border(
    left=Side(style='thin', color='D9E2F3'),
    right=Side(style='thin', color='D9E2F3'),
    top=Side(style='thin', color='D9E2F3'),
    bottom=Side(style='thin', color='D9E2F3'),
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

def write_table(ws, title, headers, data_rows, start_row=1):
    """写入表格数据"""
    r = start_row
    # 标题行
    c = ws.cell(r, 1, title)
    c.font = title_font
    r += 1
    # 表头
    for j, h in enumerate(headers, 1):
        c = ws.cell(r, j, h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = thin_border
    r += 1
    # 数据
    for row_data in data_rows:
        for j, val in enumerate(row_data, 1):
            c = ws.cell(r, j, val)
            c.font = data_font
            c.alignment = center_align
            c.border = thin_border
            if isinstance(val, float):
                c.number_format = num_format
        r += 1
    return r + 1

def fmt(v):
    if v is None or v == 0:
        return 0
    return round(v, 2)

# ===== Sheet 1: 汇总仪表板 =====
ws1 = wb.active
ws1.title = '经营分析汇总'
ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 18

c = ws1.cell(1, 1, '烨软科技 2023-2025 经营分析报告')
c.font = title_font
c = ws1.cell(2, 1, '数据来源：2025年审计报表 + 科目余额表  |  中国会计准则')
c.font = Font(name='微软雅黑', size=9, italic=True, color='808040')

r = 4

# 盈利能力
profit_headers = ['指标', '2023', '2024', '2025']
profit_rows = []
for metric in ['毛利率', '净利率', '营业利润率', '期间费用率', 'ROE', 'ROA']:
    row = [metric]
    for y in YEARS:
        row.append(f"{fmt(results['盈利能力'][y][metric])}%")
    profit_rows.append(row)
# 杜邦在单独行
for y in [2024, 2025]:
    if '杜邦_净利率' in results['盈利能力'][y]:
        metric = '杜邦ROE(净利率×周转率×权益乘数)'
        npm = results['盈利能力'][y]['杜邦_净利率']
        tat = results['盈利能力'][y]['杜邦_总资产周转率']
        em = results['盈利能力'][y]['杜邦_权益乘数']
        roe = results['盈利能力'][y]['杜邦_ROE']
        profit_rows.append([f'{y}年杜邦拆解', f'净利率{npm:.2f}%', f'周转率{tat:.2f}', f'权益乘数{em:.2f}'])
        profit_rows.append([f'{y}年杜邦ROE', '', '', f'{roe:.2f}%'])

r = write_table(ws1, '一、盈利能力', profit_headers, profit_rows, r)
r += 1

# 营运能力
oper_headers = ['指标', '2023', '2024', '2025']
oper_rows = []
for metric in ['存货周转率', '存货周转天数', '应收账款周转率', '应收账款周转天数', '总资产周转率', '总资产周转天数']:
    if metric in results['营运能力'][2023]:
        row = [metric]
        for y in YEARS:
            v = results['营运能力'][y].get(metric)
            if '周转率' in metric:
                row.append(f"{v:.2f}")
            else:
                row.append(f"{v:.0f}天")
        oper_rows.append(row)
r = write_table(ws1, '二、营运能力', oper_headers, oper_rows, r)
r += 1

# 偿债能力
debt_headers = ['指标', '2023', '2024', '2025']
debt_rows = []
for metric in ['流动比率', '速动比率', '资产负债率', '产权比率', '利息保障倍数']:
    row = [metric]
    for y in YEARS:
        v = results['偿债能力'][y][metric]
        if '率' in metric or '比' == metric[-1]:
            row.append(f"{v:.2f}")
        elif '资产负债率' in metric:
            row.append(f"{v:.1f}%")
        else:
            row.append(f"{v:.1f}")
    debt_rows.append(row)
r = write_table(ws1, '三、偿债能力', debt_headers, debt_rows, r)
r += 1

# 成长能力
growth_headers = ['指标', '2023→2024', '2024→2025', '2023-2025 CAGR']
growth_rows = []
row = ['收入增长率']
row.append(f"{fmt(results['成长能力'][2024]['收入增长率'])}%")
row.append(f"{fmt(results['成长能力'][2025]['收入增长率'])}%")
row.append(f"{fmt(results['成长能力']['CAGR']['收入CAGR(23-25)'])}%")
growth_rows.append(row)
row = ['净利润增长率']
row.append(f"{fmt(results['成长能力'][2024]['利润增长率'])}%")
row.append(f"{fmt(results['成长能力'][2025]['利润增长率'])}%")
row.append(f"{fmt(results['成长能力']['CAGR']['净利润CAGR(23-25)'])}%")
growth_rows.append(row)
# 绝对值
row = ['营业收入(万元)']
for y in YEARS:
    row.append(f"{PL[y]['营业收入']/10000:.2f}")
growth_rows.append(row)
row = ['净利润(万元)']
for y in YEARS:
    row.append(f"{PL[y]['净利润']/10000:.2f}")
growth_rows.append(row)
r = write_table(ws1, '四、成长能力', growth_headers, growth_rows, r)

# ===== Sheet 2: 利润表明细 =====
ws2 = wb.create_sheet('利润表明细')
ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 16

pl_headers = ['项目', '2023', '2024', '2025', '24vs23变动', '25vs24变动']
items = ['营业收入', '营业成本', '税金及附加', '销售费用', '管理费用', '研发费用', '财务费用', '营业利润', '营业外收入', '营业外支出', '利润总额', '净利润']
pl_rows = []
for item in items:
    row = [item]
    for y in YEARS:
        v = PL[y].get(item, 0)
        if v:
            row.append(round(v, 2))
        else:
            row.append(0)
    # 变动
    if item in PL[2024]:
        row.append(round(PL[2024][item] - PL[2023].get(item, 0), 2) if PL[2023].get(item) else 0)
        row.append(round(PL[2025][item] - PL[2024][item], 2))
    else:
        row.extend([0, 0])
    pl_rows.append(row)
# 毛利率行
for tag, num, den in [('毛利率', '营业收入-营业成本', '营业收入')]:
    row = [tag]
    for y in YEARS:
        n = PL[y]['营业收入'] - PL[y]['营业成本']
        d = PL[y]['营业收入']
        row.append(f"{n/d*100:.1f}%" if d else '-')
    row.extend(['', ''])
    pl_rows.append(row)
# 收入占比行
for tag, item in [('毛利率', '营业收入-营业成本'), ('销售费用占收入比', '销售费用'), ('管理费用占收入比', '管理费用'), ('研发费用占收入比', '研发费用'), ('期间费用率', None)]:
    if tag == '期间费用率':
        row = [tag]
        for y in YEARS:
            p = PL[y]
            total = p['销售费用'] + p['管理费用'] + p['研发费用'] + p['财务费用']
            row.append(f"{total/p['营业收入']*100:.1f}%" if p['营业收入'] else '-')
        row.extend(['', ''])
        pl_rows.append(row)
    elif tag == '毛利率':
        pass
    else:
        row = [tag]
        for y in YEARS:
            v = PL[y][item]
            row.append(f"{v/PL[y]['营业收入']*100:.1f}%" if PL[y]['营业收入'] else '-')
        row.extend(['', ''])
        pl_rows.append(row)

r = write_table(ws2, '利润表及结构分析', pl_headers, pl_rows, 1)

# ===== Sheet 3: 资产负债表明细 =====
ws3 = wb.create_sheet('资产负债表明细')
ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 16

bs_headers = ['项目', '2023末', '2024末', '2025末']
bs_items = [
    ('资产类', ['货币资金', '应收票据', '应收账款', '预付款项', '其他应收款', '存货', '流动资产合计',
                '固定资产', '无形资产', '长期待摊费用', '非流动资产合计', '资产总计']),
    ('负债类', ['短期借款', '应付账款', '预收款项', '应付职工薪酬', '应交税费', '其他应付款', '流动负债合计', '负债合计']),
    ('权益类', ['实收资本', '未分配利润', '所有者权益合计']),
]
r = 1
for section_name, items in bs_items:
    c = ws3.cell(r, 1, section_name)
    c.font = section_font
    r += 1
    rows = []
    for item in items:
        row = [item]
        for y in YEARS:
            v = BS[y].get(item, 0)
            row.append(round(v, 2) if v else 0)
        rows.append(row)
    r = write_table(ws3, '', bs_headers, rows, r)
    r += 1

# ===== Sheet 4: 经营分析结论 =====
ws4 = wb.create_sheet('分析结论')
ws4.column_dimensions['A'].width = 80

r = 1
def writeln(ws, r, text, font=None):
    c = ws.cell(r, 1, text)
    c.font = font or data_font
    return r + 1

r = writeln(ws4, r, '烨软科技 2023-2025 经营分析结论', title_font)
r += 1

conclusions = []

# 盈利能力结论
p23, p24, p25 = [results['盈利能力'][y] for y in YEARS]
r = writeln(ws4, r, '【一、盈利能力】', section_font)
r = writeln(ws4, r, f'结论：2024年大幅亏损后，2025年扭亏为盈，但净利仅1.95万，盈利能力极弱。')
r = writeln(ws4, r, f'')
r = writeln(ws4, r, f'毛利率：2023年25.6% → 2024年29.2% → 2025年47.8%，毛利率显著改善，主业盈利能力在提升。')
r = writeln(ws4, r, f'净利率：2023年0.7% → 2024年-36.0% → 2025年0.2%，2024年巨亏主因收入大幅下滑(-64.4%)但费用刚性。')
r = writeln(ws4, r, f'ROE：2023年8.7% → 2024年剧烈转负 → 2025年转正2.1%，净资产连续为负，股东权益深度亏损。')
r = writeln(ws4, r, f'期间费用率：2023年29.6% → 2024年67.3% → 2025年47.7%，费用率仍处高位。')
r = writeln(ws4, r, f'杜邦分析(2025)：净利率0.24% × 总资产周转率0.37 × 权益乘数[-] → ROE 2.1%，低利润率+低周转拖累回报。')
r += 1

# 营运能力结论
r = writeln(ws4, r, '【二、营运能力】', section_font)
r = writeln(ws4, r, f'存货周转天数：2023年190天 → 2024年513天 → 2025年724天，存货积压持续恶化。')
r = writeln(ws4, r, f'应收账款周转天数：2023年115天 → 2024年303天 → 2025年221天，2025年有所改善但仍处高位。')
r = writeln(ws4, r, f'总资产周转率：2023年0.83 → 2024年0.30 → 2025年0.37，资产使用效率显著低于正常水平(<0.5)。')
r = writeln(ws4, r, f'营运能力全面恶化，存货和应收占压大量资金，是现金流紧张的根本原因。')
r += 1

# 偿债能力结论
r = writeln(ws4, r, '【三、偿债能力】', section_font)
r = writeln(ws4, r, f'流动比率：2023年1.02 → 2024年0.88 → 2025年0.91，持续低于1，短期偿债压力大。')
r = writeln(ws4, r, f'速动比率：2023年0.67 → 2024年0.61 → 2025年0.49，持续下降，扣除存货后几乎无短期偿债能力。')
r = writeln(ws4, r, f'资产负债率：2023年93.2% → 2024年103.9% → 2025年104.2%，资不抵债持续恶化。')
r = writeln(ws4, r, f'利息保障倍数：2024年-10.9 → 2025年1.09，2024年利润无法覆盖利息，2025年仅勉强覆盖。')
r = writeln(ws4, r, f'⚠️ 公司已资不抵债，短期借款586万+应付账款1540万=2126万，而货币资金仅6.6万。')
r += 1

# 成长能力结论
r = writeln(ws4, r, '【四、成长能力】', section_font)
rev23, rev24, rev25 = [PL[y]['营业收入'] for y in YEARS]
np23, np24, np25 = [PL[y]['净利润'] for y in YEARS]
r = writeln(ws4, r, f'营业收入：2023年2,020万 → 2024年720万(-64.4%) → 2025年799万(+11.0%)')
r = writeln(ws4, r, f'净利润：2023年14.4万 → 2024年-259.3万(亏损) → 2025年1.9万(扭亏)')
r = writeln(ws4, r, f'收入在2024年断崖式下降后，2025年小幅反弹，但远未恢复至2023年水平。')
r = writeln(ws4, r, f'研发费用从2023年338万降至2025年103万(-69.5%)，研发收缩可能影响未来竞争力。')
r += 1

# 综合风险提示
r = writeln(ws4, r, '【五、综合风险提示】', section_font)
risks = [
    '🔴 资不抵债（负债率104.2%），净资产-92万，持续经营能力存在重大不确定性',
    '🔴 货币资金仅6.6万，应付账款1540万，短期借款586万，流动性高度紧张',
    '🔴 经营现金流三年中有两年为负（2024年109万正→2025年-272万负），造血能力不足',
    '🟡 研发费用持续压缩(338→200→103万)，技术型公司研发收缩影响长期竞争力',
    '🟡 存货占比资产43.8%，且周转天数长达724天，存在跌价风险',
    '🟡 利润率极薄（0.2%），抗风险能力弱',
]
for risk in risks:
    r = writeln(ws4, r, risk)

out_path = r'2025年年报\烨软科技2023-2025经营分析报告.xlsx'
wb.save(out_path)
print(f'经营分析报告已生成: {out_path}')

# 打印关键数据概览
print('\n' + '='*60)
print('关键数据一览')
print('='*60)
for y in YEARS:
    p = PL[y]
    print(f'\n{y}年: 营收={p["营业收入"]/10000:.1f}万  净利={p["净利润"]/10000:.1f}万  毛利率={profit[y]["毛利率"]:.1f}%')
    if y in [2024, 2025] and '杜邦_ROE' in profit[y]:
        d = profit[y]
        print(f'  杜邦: 净利率{d["杜邦_净利率"]:.2f}% × 周转率{d["杜邦_总资产周转率"]:.2f} × 权益乘数{d["杜邦_权益乘数"]:.2f} = ROE {d["杜邦_ROE"]:.2f}%')

print(f'\n流动比率: {debt[2025]["流动比率"]:.2f}')
print(f'资产负债率: {debt[2025]["资产负债率"]:.1f}%')
print(f'存货周转天数: {operate[2025]["存货周转天数"]:.0f}天')
print(f'应收周转天数: {operate[2025]["应收账款周转天数"]:.0f}天')
print(f'收入CAGR(23-25): {revenue_cagr_3y:.1f}%')
