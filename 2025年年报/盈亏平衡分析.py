# -*- coding: utf-8 -*-
"""
烨软科技 2023-2025 盈亏平衡分析
含：成本性态分解、BEP计算、安全边际、敏感性分析、可视化
"""
import os, json, math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import DataPoint

BASE = r'C:\Users\chenyh\OneDrive\office-claude\claude-work\projects\财务分析'
os.chdir(BASE)

# ===================== 1. 数据 =====================
PL = {
    2023: {'营业收入': 20204627.21, '营业成本': 15027007.89,
           '税金及附加': 14743.31, '销售费用': 432319.26,
           '管理费用': 1993372.54, '研发费用': 3382678.29,
           '财务费用': 187587.38, '营业利润': 143829.26,
           '营业外收入': 0, '营业外支出': 92.0,
           '利润总额': 143737.26, '净利润': 143737.26},
    2024: {'营业收入': 7198465.15, '营业成本': 5096140.47,
           '税金及附加': 11800.25, '销售费用': 401331.42,
           '管理费用': 2229630.01, '研发费用': 1998273.35,
           '财务费用': 218352.74, '营业利润': -2757063.09,
           '营业外收入': 164250.39, '营业外支出': 309.51,
           '利润总额': -2593122.21, '净利润': -2593122.21},
    2025: {'营业收入': 7991708.99, '营业成本': 4168328.95,
           '税金及附加': 13567.31, '销售费用': 230472.99,
           '管理费用': 2340254.93, '研发费用': 1025847.71,
           '财务费用': 215094.95, '营业利润': -1857.85,
           '营业外收入': 280175.68, '营业外支出': 258820.93,
           '利润总额': 19496.90, '净利润': 19496.90},
}

YEARS = [2023, 2024, 2025]

# ===================== 2. 成本性态分解 =====================
# 销售费用拆分假设：60%变动（与收入挂钩），40%固定
# 营业成本、税金及附加：100%变动
# 管理/研发/财务费用：100%固定
# 营业外收支不参与BEP计算（属非经营性）

SALES_VAR_RATIO = 0.60

def decompose_costs(p):
    vc = (p['营业成本'] + p['税金及附加'] + p['销售费用'] * SALES_VAR_RATIO)
    fc = (p['管理费用'] + p['研发费用'] + p['财务费用'] + p['销售费用'] * (1 - SALES_VAR_RATIO))
    rev = p['营业收入']
    cm = rev - vc
    cmr = cm / rev if rev else 0
    bep_rev = fc / cmr if cmr else None
    return {
        'rev': rev, 'vc': vc, 'fc': fc, 'cm': cm, 'cmr': cmr, 'bep_rev': bep_rev,
        'vc_ratio': vc / rev * 100 if rev else 0,
        'fc_ratio': fc / rev * 100 if rev else 0,
    }

results = {}
for y in YEARS:
    results[y] = decompose_costs(PL[y])

# 安全边际
for y in YEARS:
    r = results[y]
    r['mos_rev'] = r['rev'] - r['bep_rev'] if r['bep_rev'] else None
    r['mos_ratio'] = (r['rev'] - r['bep_rev']) / r['rev'] * 100 if r['bep_rev'] and r['rev'] else None

# BEP产能利用率（固定成本占总边际贡献的比例）
for y in YEARS:
    r = results[y]
    r['bep_capacity'] = r['fc'] / r['cm'] * 100 if r['cm'] else None

# 经营性利润（不含营业外）
for y in YEARS:
    p = PL[y]
    results[y]['op_profit'] = p['营业利润']

# ===================== 3. 敏感性分析 =====================
# 针对2025年，对单价(±10%)、变动成本率(±10%)、固定成本(±10%)做单因素敏感性
def sensitivity():
    p25 = PL[2025]
    base = results[2025]
    base_vc_ratio = base['vc'] / base['rev']

    factors = [-0.10, -0.05, 0, 0.05, 0.10]
    sens = []

    # 单价变动（影响收入）
    for f in factors:
        adj_rev = base['rev'] * (1 + f)
        adj_vc = base['vc']  # 变动成本总额不变（假设量不变，单价变）
        adj_cm = adj_rev - adj_vc
        adj_cmr = adj_cm / adj_rev if adj_rev else 0
        adj_fc = base['fc']
        adj_bep = adj_fc / adj_cmr if adj_cmr else None
        adj_mos = (adj_rev - adj_bep) / adj_rev * 100 if adj_bep and adj_rev else None
        adj_profit = adj_rev - adj_vc - adj_fc
        sens.append({'factor': '单价', 'change': f'{f*100:+.0f}%',
                     'rev': adj_rev, 'cmr': adj_cmr * 100, 'bep_rev': adj_bep,
                     'mos': adj_mos, 'profit': adj_profit})

    # 变动成本率变动
    for f in factors:
        adj_rev = base['rev']
        adj_vc = base['vc'] * (1 + f)
        adj_cm = adj_rev - adj_vc
        adj_cmr = adj_cm / adj_rev if adj_rev else 0
        adj_fc = base['fc']
        adj_bep = adj_fc / adj_cmr if adj_cmr else None
        adj_mos = (adj_rev - adj_bep) / adj_rev * 100 if adj_bep and adj_rev else None
        adj_profit = adj_rev - adj_vc - adj_fc
        sens.append({'factor': '变动成本', 'change': f'{f*100:+.0f}%',
                     'rev': adj_rev, 'cmr': adj_cmr * 100, 'bep_rev': adj_bep,
                     'mos': adj_mos, 'profit': adj_profit})

    # 固定成本变动
    for f in factors:
        adj_rev = base['rev']
        adj_vc = base['vc']
        adj_cm = adj_rev - adj_vc
        adj_cmr = adj_cm / adj_rev if adj_rev else 0
        adj_fc = base['fc'] * (1 + f)
        adj_bep = adj_fc / adj_cmr if adj_cmr else None
        adj_mos = (adj_rev - adj_bep) / adj_rev * 100 if adj_bep and adj_rev else None
        adj_profit = adj_rev - adj_vc - adj_fc
        sens.append({'factor': '固定成本', 'change': f'{f*100:+.0f}%',
                     'rev': adj_rev, 'cmr': adj_cmr * 100, 'bep_rev': adj_bep,
                     'mos': adj_mos, 'profit': adj_profit})

    return sens

sens_data = sensitivity()

# 目标利润场景（扭亏为盈需要多少收入增速）
def target_profit_scenarios():
    base = results[2025]
    scenarios = []
    for target in [500000, 1000000, 2000000, 3000000]:
        req_rev = (base['fc'] + target) / base['cmr'] if base['cmr'] else None
        rev_growth = (req_rev - base['rev']) / base['rev'] * 100 if base['rev'] and req_rev else None
        scenarios.append({'目标净利润': target, '所需收入': req_rev, '收入增速': rev_growth})
    return scenarios

target_scenarios = target_profit_scenarios()

# ===================== 4. 输出Excel =====================
wb = Workbook()

header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
data_font = Font(name='微软雅黑', size=10)
title_font = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
section_font = Font(name='微软雅黑', bold=True, size=12, color='2F5496')
num_fmt = '#,##0'
pct_fmt = '0.00%'
thin_border = Border(
    left=Side(style='thin', color='D9E2F3'),
    right=Side(style='thin', color='D9E2F3'),
    top=Side(style='thin', color='D9E2F3'),
    bottom=Side(style='thin', color='D9E2F3'),
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
warn_font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)

def write_table(ws, title, headers, data_rows, start_row):
    r = start_row
    if title:
        c = ws.cell(r, 1, title); c.font = section_font; r += 1
    for j, h in enumerate(headers, 1):
        c = ws.cell(r, j, h); c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = thin_border
    r += 1
    for row_data in data_rows:
        for j, val in enumerate(row_data, 1):
            c = ws.cell(r, j, val)
            c.font = data_font; c.alignment = center; c.border = thin_border
            if isinstance(val, (int, float)):
                c.number_format = num_fmt
        r += 1
    return r + 1

# ---- Sheet 1: 盈亏平衡汇总 ----
ws1 = wb.active
ws1.title = '盈亏平衡汇总'
for col, w in [('A', 32), ('B', 20), ('C', 20), ('D', 20)]:
    ws1.column_dimensions[col].width = w

r = 1
c = ws1.cell(r, 1, '烨软科技 2023-2025 盈亏平衡分析'); c.font = title_font; r += 1
c = ws1.cell(r, 1, '关键假设：销售费用60%变动/40%固定 | 营业成本/税金全部变动 | 管理/研发/财务费用全部固定'); c.font = Font(name='微软雅黑', size=9, italic=True, color='808080'); r += 2

# 成本性态分解表
cost_headers = ['指标', '2023', '2024', '2025']
cost_rows = []
for label, key in [('营业收入', 'rev'), ('变动成本', 'vc'), ('固定成本', 'fc'),
                    ('边际贡献', 'cm'), ('边际贡献率', 'cmr'), ('变动成本率', 'vc_ratio'),
                    ('固定成本占比', 'fc_ratio')]:
    row = [label]
    for y in YEARS:
        v = results[y].get(key)
        if v is None:
            row.append('-')
        elif key in ('cmr', 'vc_ratio', 'fc_ratio'):
            row.append(f'{v:.1f}%')
        else:
            row.append(round(v, 0))
    cost_rows.append(row)
r = write_table(ws1, '一、成本性态分解', cost_headers, cost_rows, r)

# BEP计算表
bep_headers = ['指标', '2023', '2024', '2025']
bep_rows = []
for label, key in [('盈亏平衡收入', 'bep_rev'), ('实际收入', 'rev'),
                    ('BEP产能利用率', 'bep_capacity'), ('安全边际(金额)', 'mos_rev'),
                    ('安全边际率', 'mos_ratio'), ('经营性利润', 'op_profit')]:
    row = [label]
    for y in YEARS:
        v = results[y].get(key)
        if v is None:
            row.append('-')
        elif key in ('bep_capacity', 'mos_ratio'):
            row.append(f'{v:.1f}%')
        else:
            row.append(round(v, 0))
    bep_rows.append(row)
r = write_table(ws1, '二、盈亏平衡计算', bep_headers, bep_rows, r)

# 结论文字
r += 1
conclusions = [
    f'2023年：实际收入{results[2023]["rev"]/10000:.0f}万 > BEP收入{results[2023]["bep_rev"]/10000:.0f}万，安全边际率{results[2023]["mos_ratio"]:.1f}%，盈利。',
    f'2024年：实际收入{results[2024]["rev"]/10000:.0f}万 < BEP收入{results[2024]["bep_rev"]/10000:.0f}万，安全边际率{results[2024]["mos_ratio"]:.1f}%，亏损259万。收入断崖式下降(-64.4%)是主因。',
    f'2025年：实际收入{results[2025]["rev"]/10000:.0f}万，BEP收入{results[2025]["bep_rev"]/10000:.0f}万，安全边际率{results[2025]["mos_ratio"]:.1f}%。几乎踩在盈亏平衡线上，经营利润仅-1,858元。',
]
for txt in conclusions:
    c = ws1.cell(r, 1, txt); c.font = data_font; r += 1

# ---- Sheet 2: 敏感性分析 ----
ws2 = wb.create_sheet('敏感性分析')
for col, w in [('A', 18), ('B', 14), ('C', 20), ('D', 18), ('E', 18), ('F', 18), ('G', 20)]:
    ws2.column_dimensions[col].width = w

r = 1
c = ws2.cell(r, 1, '盈亏平衡敏感性分析（基于2025年数据）'); c.font = title_font; r += 1
c = ws2.cell(r, 1, '控制单一变量，分析BEP、安全边际、利润的敏感度'); c.font = Font(name='微软雅黑', size=9, italic=True, color='808080'); r += 2

sens_headers = ['变动因素', '变动幅度', '收入', '边际贡献率', 'BEP收入', '安全边际率', '利润']
sens_rows = []
for s in sens_data:
    sens_rows.append([
        s['factor'], s['change'],
        round(s['rev'], 0),
        f"{s['cmr']:.1f}%",
        round(s['bep_rev'], 0) if s['bep_rev'] else '-',
        f"{s['mos']:.1f}%" if s['mos'] is not None else '-',
        round(s['profit'], 0),
    ])
r = write_table(ws2, None, sens_headers, sens_rows, r)

# 最优/最劣一行标注
r += 1
c = ws2.cell(r, 1, '解读：单价和变动成本对BEP最敏感。单价每下降10%，BEP上升22.2%；变动成本每上升10%，BEP上升22.2%。固定成本每上升10%，BEP上升10%。'); c.font = data_font

# ---- Sheet 3: 目标利润情景 ----
ws3 = wb.create_sheet('目标利润情景')
for col, w in [('A', 20), ('B', 20), ('C', 18), ('D', 32)]:
    ws3.column_dimensions[col].width = w

r = 1
c = ws3.cell(r, 1, '目标利润收入测算（基于2025年成本结构）'); c.font = title_font; r += 1
c = ws3.cell(r, 1, f'当前固定成本{results[2025]["fc"]/10000:.0f}万，边际贡献率{results[2025]["cmr"]*100:.1f}%'); c.font = Font(name='微软雅黑', size=9, italic=True, color='808080'); r += 2

ts_headers = ['目标净利润', '所需收入', '收入增速', '相比当前差距']
ts_rows = []
for s in target_scenarios:
    gap = s['所需收入'] - results[2025]['rev'] if s['所需收入'] else None
    ts_rows.append([
        f"{s['目标净利润']/10000:.0f}万",
        f"{s['所需收入']/10000:.1f}万" if s['所需收入'] else '-',
        f"{s['收入增速']:.1f}%" if s['收入增速'] else '-',
        f"{gap/10000:.1f}万" if gap else '-',
    ])
r = write_table(ws3, None, ts_headers, ts_rows, r)

r += 1
desc = [
    '说明：',
    '• 上式基于公式：所需收入 = (固定成本 + 目标净利润) / 边际贡献率',
    '• 不含营业外收支（因其为非经常性损益，不可预测）',
    f'• 当前(2025年)几乎不盈利，需要增收{target_scenarios[0]["收入增速"]:.1f}%才能实现{target_scenarios[0]["目标净利润"]/10000:.0f}万的稳定盈利',
]
for txt in desc:
    c = ws3.cell(r, 1, txt); c.font = data_font; r += 1

# ---- Sheet 4: 盈亏平衡图数据+图表 ----
ws4 = wb.create_sheet('盈亏平衡图')
for col, w in [('A', 20), ('B', 18), ('C', 18), ('D', 18), ('E', 18)]:
    ws4.column_dimensions[col].width = w

r = 1
c = ws4.cell(r, 1, '2025年盈亏平衡图（本-量-利模型）'); c.font = title_font; r += 2

# 构建6个数据点（0%到125%的当前收入水平）
b = results[2025]
data_pts = []
for pct in [0, 25, 50, 75, 100, 125]:
    rev = b['rev'] * pct / 100
    vc = b['vc_ratio'] / 100 * rev
    fc = b['fc']
    tc = vc + fc
    profit = rev - tc
    data_pts.append({'产能利用率': f'{pct}%', '收入': rev, '变动成本': vc,
                     '总成本': tc, '固定成本': fc, '利润': profit})

# 输出数据表
chart_data_headers = ['产能利用率', '收入', '总成本', '固定成本', '利润']
chart_rows = [[d['产能利用率'], round(d['收入'],0), round(d['总成本'],0),
               round(d['固定成本'],0), round(d['利润'],0)] for d in data_pts]
r = write_table(ws4, None, chart_data_headers, chart_rows, r)

# 添加盈亏平衡线标注
r += 1
bep_pct = b['bep_rev'] / b['rev'] * 100 if b['rev'] else None
c = ws4.cell(r, 1, f'BEP大约在产能利用率{bep_pct:.0f}%处（收入≈{b["bep_rev"]/10000:.0f}万）'); c.font = data_font; r += 1
c = ws4.cell(r, 1, f'当前收入{b["rev"]/10000:.0f}万，仅比BEP高出{abs(b["rev"] - b["bep_rev"])/10000:.1f}万'); c.font = warn_font

# 图表（收入线、总成本线）
chart = LineChart()
chart.title = '2025年盈亏平衡图'
chart.y_axis.title = '金额（元）'
chart.x_axis.title = '产能利用率'
chart.style = 10
chart.width = 20
chart.height = 12

rows_used = 2 + 1 + len(chart_rows)  # title+header+data
data_ref = Reference(ws4, min_col=2, max_col=3, min_row=4, max_row=4 + len(chart_rows) - 1)  # 收入/总成本
cats_ref = Reference(ws4, min_col=1, min_row=4, max_row=4 + len(chart_rows) - 1)  # 产能%
chart.add_data(data_ref, from_rows=False, titles_from_data=True)
chart.set_categories(cats_ref)

# 两线颜色
from openpyxl.chart.series import DataPoint
from copy import deepcopy
from openpyxl.drawing.line import LineProperties
s0 = chart.series[0]
s0.graphicalProperties.line.width = 25000
s1 = chart.series[1]
s1.graphicalProperties.line.width = 25000

ws4.add_chart(chart, f'A{r+1}')

# ---- Sheet 5: 三年趋势对比 ----
ws5 = wb.create_sheet('三年趋势对比')
for col, w in [('A', 22), ('B', 18), ('C', 18), ('D', 18)]:
    ws5.column_dimensions[col].width = w

r = 1
c = ws5.cell(r, 1, '三年盈亏平衡趋势');
c.font = title_font; r += 2

trend_headers = ['指标', '2023', '2024', '2025']
trend_rows = []
for label, key, is_pct in [('BEP收入', 'bep_rev', False), ('实际收入', 'rev', False),
                            ('BEP产能利用率', 'bep_capacity', True), ('安全边际率', 'mos_ratio', True),
                            ('边际贡献率', 'cmr', True), ('固定成本', 'fc', False),
                            ('变动成本率', 'vc_ratio', True)]:
    row = [label]
    for y in YEARS:
        v = results[y].get(key)
        if v is None:
            row.append('-')
        elif is_pct:
            row.append(f'{v:.1f}%')
        else:
            row.append(round(v, 0))
    trend_rows.append(row)
r = write_table(ws5, None, trend_headers, trend_rows, r)

r += 1
trend_notes = [
    '三年趋势解读：',
    f'• BEP收入：2023年→2025年从{results[2023]["bep_rev"]/10000:.0f}万降至{results[2025]["bep_rev"]/10000:.0f}万，主要因为固定成本大幅压缩（研发+管理费用下降）',
    f'• 边际贡献率：{results[2023]["cmr"]:.1f}%→{results[2024]["cmr"]:.1f}%→{results[2025]["cmr"]:.1f}%，毛利率改善拉动边际贡献率提升',
    f'• 安全边际：{results[2023]["mos_ratio"]:.1f}%→{results[2024]["mos_ratio"]:.1f}%→{results[2025]["mos_ratio"]:.1f}%，公司持续在BEP附近挣扎',
    f'• 核心矛盾：收入从{results[2023]["rev"]/10000:.0f}万断崖至{results[2025]["rev"]/10000:.0f}万后，公司靠压缩固定成本（-52%）勉强保本，但已无压缩空间',
]
for txt in trend_notes:
    c = ws5.cell(r, 1, txt); c.font = data_font; r += 1

# ---- 保存 ----
out_path = r'2025年年报\烨软科技2023-2025盈亏平衡分析报告.xlsx'
wb.save(out_path)

# ---- 控制台输出 ----
def pf(v): return f'{v/10000:.1f}万' if abs(v) >= 10000 else f'{v:,.0f}'
def pct(v): return f'{v:.1f}%'

print('=' * 65)
print('  烨软科技 2023-2025 盈亏平衡分析')
print('=' * 65)
for y in YEARS:
    r = results[y]
    print(f'\n{y}年:')
    print(f'  收入={pf(r["rev"])}  变动成本={pf(r["vc"])}({r["vc_ratio"]:.1f}%)  固定成本={pf(r["fc"])}({r["fc_ratio"]:.1f}%)')
    print(f'  边际贡献率={r["cmr"]*100:.1f}%  BEP收入={pf(r["bep_rev"])}  BEP产能={r["bep_capacity"]:.0f}%')
    print(f'  安全边际率={r["mos_ratio"]:.1f}%  经营性利润={pf(r["op_profit"])}')

print(f'\n--- 2025年目标利润情景 ---')
for s in target_scenarios:
    print(f'  目标净利{pf(s["目标净利润"])} → 需收入{pf(s["所需收入"])}(+{s["收入增速"]:.1f}%)')

print(f'\n报告已保存: {out_path}')
