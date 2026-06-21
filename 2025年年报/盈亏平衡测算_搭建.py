# -*- coding: utf-8 -*-
"""
Breakeven model for Yesoft Tech.
All data cells are formulas referencing source sheets — changing yellow cells triggers global update.
"""
import os, sys, json, math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.chart import LineChart, ScatterChart, Reference
from openpyxl.chart.series import SeriesLabel

BASE = r'C:\Users\chenyh\OneDrive\office-claude\claude-work\projects\财务分析'
os.chdir(BASE)

wb = Workbook()

# ========== Styles ==========
hdr_font   = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
hdr_fill   = PatternFill('solid', fgColor='2F5496')
param_fill = PatternFill('solid', fgColor='FFF2CC')
t_font     = Font(name='Microsoft YaHei', bold=True, size=14, color='2F5496')
s_font     = Font(name='Microsoft YaHei', bold=True, size=12, color='2F5496')
d_font     = Font(name='Microsoft YaHei', size=10)
b_font     = Font(name='Microsoft YaHei', bold=True, size=10)
n_font     = Font(name='Microsoft YaHei', size=9, italic=True, color='808080')
warn_f     = Font(name='Microsoft YaHei', size=10, color='FF0000', bold=True)
big_font   = Font(name='Microsoft YaHei', bold=True, size=12)
nfmt       = '#,##0'
pfmt       = '0.00%'
border     = Border(
    left=Side('thin','D9E2F3'), right=Side('thin','D9E2F3'),
    top=Side('thin','D9E2F3'),  bottom=Side('thin','D9E2F3'))
calign     = Alignment(horizontal='center', vertical='center', wrap_text=True)
gfill      = PatternFill('solid', fgColor='C6EFCE')
rfill      = PatternFill('solid', fgColor='FFC7CE')

def sc(ws, r, c, v, font=d_font, fill=None, fmt=None):
    cell = ws.cell(r, c, v)
    cell.font = font; cell.alignment = calign; cell.border = border
    if fill: cell.fill = fill
    if fmt:  cell.number_format = fmt
    return cell

def hdr(ws, r, cols):
    for j, h in enumerate(cols, 1):
        sc(ws, r, j, h, font=hdr_font, fill=hdr_fill)

# ========== Sheet names (English internally, Chinese display) ==========
SN_GUIDE = '使用说明'
SN_PARAM = '产品参数'
SN_BEP   = '盈亏平衡计算'
SN_SENS  = '敏感性分析'
SN_VIZ   = '可视化'

# ================================================================
# Sheet0: Guide
# ================================================================
ws0 = wb.active
ws0.title = SN_GUIDE
ws0.column_dimensions['A'].width = 3
ws0.column_dimensions['B'].width = 55
ws0.column_dimensions['C'].width = 6

r = 1
ws0.cell(r, 2, '烨软科技 盈亏平衡测算模型').font = t_font; r += 1
ws0.cell(r, 2, '—— 两个产品（120 / 155）要卖多少才不亏钱？').font = Font(name='Microsoft YaHei', size=12, color='666666'); r += 2

ws0.cell(r, 2, '怎么用？三步').font = Font(name='Microsoft YaHei', bold=True, size=14, color='2F5496'); r += 1
for title, desc in [
    ('第一步：点「产品参数」', '黄色格子改销量。白色格子自动算，改完回车利润自动刷新。'),
    ('第二步：点「盈亏平衡计算」', '5 BEP收入 = 保本线   7 安全边际 = 比保本线多多少（负=亏）   盈亏状态 = 赚/亏'),
    ('第三步：点「可视化」', '图A 收入vs成本交叉点 = 保本位置\n图B 利润曲线 = 啥时候开始赚钱\n图C 等利润线 = 两个产品怎么搭才保本，红圈=你现在在哪'),
]:
    ws0.cell(r, 2, title).font = Font(name='Microsoft YaHei', bold=True, size=11); r += 1
    c = ws0.cell(r, 2, desc); c.font = d_font; c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws0.row_dimensions[r].height = 42; r += 1
r += 1

ws0.cell(r, 2, '当前结果（数据来源于你填的参数）').font = Font(name='Microsoft YaHei', bold=True, size=14, color='2F5496'); r += 1
for label, val, note in [
    ('公司一年固定开销', '349 万', '工资+房租+利息等，详见「盈亏平衡计算」'),
    ('贸易业务贡献', '60 万', '固定利润，不参与本表计算'),
    ('120和155需赚回', '289 万', '349-60，这是两个产品要拼的'),
    ('保本需要卖到', '525 万', '120+155合计收入'),
    ('现在实际卖', '330 万', '假设各卖500个，去「产品参数」改'),
    ('当前结果', '亏 107 万', '差 195 万才保本'),
]:
    ws0.cell(r, 2, f'{label}：').font = Font(name='Microsoft YaHei', bold=True, size=10)
    ws0.cell(r, 3, f'{val}    {note}').font = Font(name='Microsoft YaHei', size=10, color='555555'); r += 1
r += 1

ws0.cell(r, 2, '快速算：想赚多少钱，得卖多少？').font = Font(name='Microsoft YaHei', bold=True, size=14, color='2F5496'); r += 1
ws0.cell(r, 2, '公式：(289万 + 想要的利润) / 55% = 需要卖到').font = Font(name='Microsoft YaHei', size=11); r += 1
for target, rev in [('不亏（保本）', '525 万'), ('赚 50 万', '616 万'), ('赚 100 万', '707 万')]:
    ws0.cell(r, 2, f'  {target}：需要卖到 {rev}').font = Font(name='Microsoft YaHei', size=10); r += 1
r += 1

ws0.cell(r, 2, '注意事项').font = Font(name='Microsoft YaHei', bold=True, size=14, color='2F5496'); r += 1
for n in [
    '只改黄色格子，其他格子有公式，改了就坏了',
    '毛利率会自动根据你填的单价/成本算，不用手填',
    '利息按当前贷款算（1.5万/月），还了贷款要改月利息',
    '折旧 30 万/年是非现金支出，现金流比报表利润好看',
]:
    ws0.cell(r, 2, f'· {n}').font = d_font; r += 1

# ================================================================
# Sheet1: Product Parameters
# ================================================================
ws1 = wb.create_sheet(SN_PARAM)
for cl, w in [('A', 26), ('B', 18), ('C', 18), ('D', 18), ('E', 18)]:
    ws1.column_dimensions[cl].width = w

r = 1
ws1.cell(r, 1, '烨软科技 盈亏平衡测算模型').font = t_font; r += 1
ws1.cell(r, 1, '黄色 = 手动改   白色 = 公式自动').font = n_font; r += 2

ws1.cell(r, 1, '一、产品参数').font = s_font; r += 1
hdr(ws1, r, ['', '120', '155', '研发产品', '贸易']); r += 1

# Row references (will be hardcoded as numbers for formulas)
# R6: Sales Qty
R_SALES = r
sc(ws1, r, 1, '年销量（个）')
sc(ws1, r, 2, 500,  font=b_font, fill=param_fill, fmt=nfmt)
sc(ws1, r, 3, 500,  font=b_font, fill=param_fill, fmt=nfmt)
sc(ws1, r, 4, 1,    font=b_font, fill=param_fill, fmt=nfmt)
sc(ws1, r, 5, 1,    font=b_font, fill=param_fill, fmt=nfmt); r += 1

# R7: Price
R_PRICE = r
sc(ws1, r, 1, '单价（元）')
sc(ws1, r, 2, 2600,    font=b_font, fill=param_fill, fmt=nfmt)
sc(ws1, r, 3, 4000,    font=b_font, fill=param_fill, fmt=nfmt)
sc(ws1, r, 4, 1000000, font=b_font, fill=param_fill, fmt=nfmt)
sc(ws1, r, 5, 3000000, font=b_font, fill=param_fill, fmt=nfmt); r += 1

# R8: Unit Cost
R_UCOST = r
sc(ws1, r, 1, '单位材料成本（元）')
sc(ws1, r, 2, 1170,    font=b_font, fill=param_fill, fmt=nfmt)
sc(ws1, r, 3, 1800,    font=b_font, fill=param_fill, fmt=nfmt)
sc(ws1, r, 4, 1000000, font=b_font, fill=param_fill, fmt=nfmt)
sc(ws1, r, 5, 2400000, font=b_font, fill=param_fill, fmt=nfmt); r += 1

r += 1  # blank

# R10: Unit Margin = Price - UnitCost
R_UMARGIN = r
sc(ws1, r, 1, '单位毛利（元）')
for c in range(2, 6):
    cl = get_column_letter(c)
    sc(ws1, r, c, f'={cl}{R_PRICE}-{cl}{R_UCOST}', fmt=nfmt)
r += 1

# R11: Gross Margin% = UnitMargin / Price
R_GMRATIO = r
sc(ws1, r, 1, '毛利率')
for c in range(2, 6):
    cl = get_column_letter(c)
    sc(ws1, r, c, f'=IF({cl}{R_PRICE}=0,"-",{cl}{R_UMARGIN}/{cl}{R_PRICE})', fmt=pfmt)
r += 1

r += 1  # blank

# R13: Annual Revenue = Sales * Price
R_REVENUE = r
sc(ws1, r, 1, '年收入')
for c in range(2, 6):
    cl = get_column_letter(c)
    sc(ws1, r, c, f'={cl}{R_SALES}*{cl}{R_PRICE}', fmt=nfmt)
r += 1

# R14: Annual Gross Margin = UnitMargin * Sales
R_GMAMT = r
sc(ws1, r, 1, '年毛利额')
for c in range(2, 6):
    cl = get_column_letter(c)
    sc(ws1, r, c, f'={cl}{R_UMARGIN}*{cl}{R_SALES}', fmt=nfmt)
r += 1

# --- Summary ---
r += 1  # R16
ws1.cell(r, 1, '二、汇总').font = s_font; r += 1

R_TOTAL_REV = r  # R17
sc(ws1, r, 1, '总收入', font=b_font)
sc(ws1, r, 2, f'=SUM(B{R_REVENUE}:E{R_REVENUE})', font=b_font, fmt=nfmt); r += 1

R_TOTAL_GM = r  # R18
sc(ws1, r, 1, '总毛利', font=b_font)
sc(ws1, r, 2, f'=SUM(B{R_GMAMT}:E{R_GMAMT})', font=b_font, fmt=nfmt); r += 1

R_COMP_GM = r  # R19
sc(ws1, r, 1, '综合毛利率', font=b_font)
sc(ws1, r, 2, f'=IF(B{R_TOTAL_REV}=0,"-",B{R_TOTAL_GM}/B{R_TOTAL_REV})', font=b_font, fmt=pfmt); r += 1

R_VAR_REV = r  # R20: 120&155 total revenue
sc(ws1, r, 1, '120&155 合计收入', font=b_font)
sc(ws1, r, 2, f'=B{R_REVENUE}+C{R_REVENUE}', font=b_font, fmt=nfmt); r += 1

R_CMR = r  # R21: 120&155 combined margin = (GM120+GM155)/(Rev120+Rev155)
sc(ws1, r, 1, '120&155 毛利率', font=b_font)
sc(ws1, r, 2, f'=IF(B{R_VAR_REV}=0,"-",(B{R_GMAMT}+C{R_GMAMT})/B{R_VAR_REV})', font=b_font, fmt=pfmt)
r += 2

# --- Quick profit check ---
ws1.cell(r, 1, '三、利润快算').font = s_font; r += 1

R_P_GM = r  # R24
sc(ws1, r, 1, '总毛利', font=b_font)
sc(ws1, r, 2, f'=B{R_TOTAL_GM}', font=b_font, fmt=nfmt); r += 1

R_P_FC = r  # R25 — will be set to reference Sheet2 after Sheet2 is built; placeholder
sc(ws1, r, 1, '固定费用', font=b_font)
sc(ws1, r, 2, 0, font=b_font, fmt=nfmt)
R25_TEMP = r
r += 1

R_P_PROFIT = r  # R26
sc(ws1, r, 1, '利润', font=big_font)
sc(ws1, r, 2, f'=B{R_P_GM}-B{R_P_FC}', font=big_font, fmt=nfmt)

# ================================================================
# Sheet2: BEP Calculation
# ================================================================
ws2 = wb.create_sheet(SN_BEP)
for cl, w in [('A', 36), ('B', 24), ('C', 24)]:
    ws2.column_dimensions[cl].width = w

r = 1
ws2.cell(r, 1, '盈亏平衡计算').font = t_font; r += 2

# --- Fixed costs ---
ws2.cell(r, 1, '一、固定费用明细（年）').font = s_font; r += 1
hdr(ws2, r, ['项目', '月（可调）', '年（=月x12）']); r += 1

FC_START = r
fc_items = [
    ('人工（含社保）', 230000),
    ('房租', 14036),
    ('水电', 1500),
    ('物管', 1750),
    ('折旧', 25000),
    ('办公费', 2000),
    ('中介费', 1500),
    ('利息', 15000),
]
for name, monthly in fc_items:
    sc(ws2, r, 1, name)
    sc(ws2, r, 2, monthly, font=b_font, fill=param_fill, fmt=nfmt)  # yellow=editable
    sc(ws2, r, 3, f'=B{r}*12', fmt=nfmt)
    r += 1
FC_END = r - 1

FC_TOTAL_ROW = r
sc(ws2, r, 1, '固定费用合计', font=b_font)
sc(ws2, r, 2, '-', fmt=nfmt)
sc(ws2, r, 3, f'=SUM(C{FC_START}:C{FC_END})', font=b_font, fmt=nfmt)
r += 2

# --- BEP Calculation ---
ws2.cell(r, 1, '二、盈亏平衡计算').font = s_font; r += 1
BEP_START = r

# 1 FC total
sc(ws2, r, 1, '1 固定费用合计')
sc(ws2, r, 2, f'=C{FC_TOTAL_ROW}', font=b_font, fmt=nfmt); r += 1

# 2 Trade GM = PARAM!E14
sc(ws2, r, 1, '2 贸易年毛利（=产品参数!E列年毛利额）')
sc(ws2, r, 2, f"='{SN_PARAM}'!E{R_GMAMT}", fmt=nfmt); r += 1

# 3 Gap = 1 - 2
BEP_GAP_ROW = r
sc(ws2, r, 1, '3 = 1 - 2  需120&155弥补的毛利缺口')
sc(ws2, r, 2, f'=B{r-2}-B{r-1}', font=b_font, fmt=nfmt); r += 1

# 4 CMR = PARAM!B21
BEP_CMR_ROW = r
sc(ws2, r, 1, '4 120&155 毛利率（=产品参数!B21）')
sc(ws2, r, 2, f"='{SN_PARAM}'!B{R_CMR}", fmt=pfmt); r += 1

# 5 BEP Rev = 3 / 4
BEP_REV_ROW = r
sc(ws2, r, 1, '5 = 3 / 4  BEP收入（120&155需达到）')
sc(ws2, r, 2, f'=IF(B{r-1}=0,"-",B{r-2}/B{r-1})', font=b_font, fmt=nfmt); r += 1

r += 1  # blank

# 6 Actual = PARAM!B20
BEP_ACTUAL_ROW = r
sc(ws2, r, 1, '6 当前120&155实际收入（=产品参数!B20）')
sc(ws2, r, 2, f"='{SN_PARAM}'!B{R_VAR_REV}", fmt=nfmt); r += 1

# 7 Margin of Safety
BEP_MOS_AMT_ROW = r
sc(ws2, r, 1, '7 = 6 - 5  安全边际（金额）')
sc(ws2, r, 2, f'=B{r-1}-B{BEP_REV_ROW}', font=b_font, fmt=nfmt); r += 1

# 8 MOS%
BEP_MOS_PCT_ROW = r
sc(ws2, r, 1, '8 = 7 / 6  安全边际率')
sc(ws2, r, 2, f'=IF(B{r-2}=0,"-",B{r-1}/B{r-2})', font=b_font, fmt=pfmt); r += 1

r += 1

# P&L status
sc(ws2, r, 1, '-> 盈亏状态')
sc(ws2, r, 2, f'=IF(B{BEP_MOS_AMT_ROW}>=0,"[ 盈 利 ]","[ 亏 损 ]")', font=warn_f)
r += 2

# --- BEP decomposition ---
ws2.cell(r, 1, '三、BEP组合分解（任意组合，120/155收入合计>=BEP收入即可保本）').font = s_font; r += 1
hdr(ws2, r, ['情景', '120年销量', '155年销量', '120收入', '155收入', '120&155合计', '是否保本']); r += 1

# Pure 120
sc(ws2, r, 1, '纯卖120')
sc(ws2, r, 2, f"=ROUNDUP(B{BEP_REV_ROW}/'{SN_PARAM}'!B{R_PRICE},0)", fmt=nfmt)
sc(ws2, r, 3, 0, fmt=nfmt)
sc(ws2, r, 4, f"=B{r}*'{SN_PARAM}'!B{R_PRICE}", fmt=nfmt)
sc(ws2, r, 5, 0, fmt=nfmt)
sc(ws2, r, 6, f'=D{r}+E{r}', font=b_font, fmt=nfmt)
sc(ws2, r, 7, f'=IF(F{r}>=B{BEP_REV_ROW},"OK 保本","X 亏损")'); r += 1

# Pure 155
sc(ws2, r, 1, '纯卖155')
sc(ws2, r, 2, 0, fmt=nfmt)
sc(ws2, r, 3, f"=ROUNDUP(B{BEP_REV_ROW}/'{SN_PARAM}'!C{R_PRICE},0)", fmt=nfmt)
sc(ws2, r, 4, 0, fmt=nfmt)
sc(ws2, r, 5, f"=C{r}*'{SN_PARAM}'!C{R_PRICE}", fmt=nfmt)
sc(ws2, r, 6, f'=D{r}+E{r}', font=b_font, fmt=nfmt)
sc(ws2, r, 7, f'=IF(F{r}>=B{BEP_REV_ROW},"OK 保本","X 亏损")'); r += 1

# Current
sc(ws2, r, 1, '当前组合')
sc(ws2, r, 2, f"='{SN_PARAM}'!B{R_SALES}", fmt=nfmt)
sc(ws2, r, 3, f"='{SN_PARAM}'!C{R_SALES}", fmt=nfmt)
sc(ws2, r, 4, f"=B{r}*'{SN_PARAM}'!B{R_PRICE}", fmt=nfmt)
sc(ws2, r, 5, f"=C{r}*'{SN_PARAM}'!C{R_PRICE}", fmt=nfmt)
sc(ws2, r, 6, f'=D{r}+E{r}', font=b_font, fmt=nfmt)
sc(ws2, r, 7, f'=IF(F{r}>=B{BEP_REV_ROW},"OK 保本","X 亏损")')
r += 2

# --- Current profit ---
ws2.cell(r, 1, '四、当前利润').font = s_font; r += 1
sc(ws2, r, 1, '总毛利（=产品参数!B18）')
sc(ws2, r, 2, f"='{SN_PARAM}'!B{R_TOTAL_GM}", font=b_font, fmt=nfmt); r += 1
sc(ws2, r, 1, '固定费用（=上面合计）')
sc(ws2, r, 2, f'=C{FC_TOTAL_ROW}', font=b_font, fmt=nfmt); r += 1
sc(ws2, r, 1, '利润')
sc(ws2, r, 2, f'=B{r-2}-B{r-1}', font=big_font, fmt=nfmt)

# ===== Fix Sheet1 R25 reference =====
ws1.cell(R25_TEMP, 2).value = f"='{SN_BEP}'!C{FC_TOTAL_ROW}"

# ================================================================
# Sheet3: Sensitivity Matrix (all formulas + conditional formatting)
# ================================================================
ws3 = wb.create_sheet(SN_SENS)
for cl, w in [('A', 15)] + [(get_column_letter(i), 14) for i in range(2, 14)]:
    ws3.column_dimensions[cl].width = w

r = 1
ws3.cell(r, 1, '产品组合敏感性分析（利润矩阵）').font = t_font; r += 1
ws3.cell(r, 1, '行=120年销量   列=155年销量   格=利润(元)   绿盈红亏  (公式联动，改参数自动刷新)').font = n_font; r += 2

q120_range = list(range(0, 2501, 250))
q155_range = list(range(0, 2001, 200))

MX_HEADER_ROW = r
sc(ws3, r, 1, '120 / 155 ->', font=b_font)
for j, q155 in enumerate(q155_range):
    sc(ws3, r, j+2, q155, font=hdr_font, fill=hdr_fill)
r += 1
MX_DATA_START = r

for q120 in q120_range:
    sc(ws3, r, 1, q120, font=b_font)
    for j, q155 in enumerate(q155_range):
        cl_top = get_column_letter(j+2)
        # Profit = Q120*UM120 + Q155*UM155 + TradeGM - FC
        formula = (
            f"=A{r}*'{SN_PARAM}'!B{R_UMARGIN}"
            f"+{cl_top}${MX_HEADER_ROW}*'{SN_PARAM}'!C{R_UMARGIN}"
            f"+'{SN_PARAM}'!E{R_GMAMT}"
            f"-'{SN_BEP}'!C{FC_TOTAL_ROW}"
        )
        sc(ws3, r, j+2, formula, font=Font(name='Microsoft YaHei', size=9), fmt=nfmt)
    r += 1
MX_DATA_END = r - 1

# Conditional formatting
min_cl = get_column_letter(2)
max_cl = get_column_letter(len(q155_range)+1)
mx_range = f'{min_cl}{MX_DATA_START}:{max_cl}{MX_DATA_END}'

green_ds = DifferentialStyle(fill=gfill)
red_ds   = DifferentialStyle(fill=rfill)

ws3.conditional_formatting.add(mx_range,
    Rule(type='cellIs', operator='greaterThan', formula=['0'], dxf=green_ds))
ws3.conditional_formatting.add(mx_range,
    Rule(type='cellIs', operator='lessThanOrEqual', formula=['0'], dxf=red_ds))

# Legend & current
r = MX_DATA_END + 2
sc(ws3, r, 1, '图例:', font=b_font)
sc(ws3, r, 2, '绿色=盈利', font=d_font, fill=gfill); r += 1
sc(ws3, r, 2, '红色=亏损', font=d_font, fill=rfill); r += 2

sc(ws3, r, 1, '当前利润')
sc(ws3, r, 2, f"='{SN_PARAM}'!B{R_TOTAL_GM}-'{SN_BEP}'!C{FC_TOTAL_ROW}", font=b_font, fmt=nfmt); r += 1
sc(ws3, r, 1, '当前120销量')
sc(ws3, r, 2, f"='{SN_PARAM}'!B{R_SALES}", fmt=nfmt); r += 1
sc(ws3, r, 1, '当前155销量')
sc(ws3, r, 2, f"='{SN_PARAM}'!C{R_SALES}", fmt=nfmt)

# ================================================================
# Sheet4: Charts
# ================================================================
ws4 = wb.create_sheet(SN_VIZ)
for cl, w in [('A', 24), ('B', 20), ('C', 20), ('D', 20), ('E', 20), ('F', 20), ('G', 20), ('H', 20), ('I', 20)]:
    ws4.column_dimensions[cl].width = w

r = 1
ws4.cell(r, 1, '盈亏平衡可视化（A.本量利  B.利润曲线  C.等利润线）').font = t_font; r += 1
ws4.cell(r, 1, '* 数据全部公式引用，改Sheet1/Sheet2 -> 图表自动更新').font = n_font; r += 2

# ==== Chart A: CVP ====
ws4.cell(r, 1, '图A: 本量利图数据').font = s_font; r += 1

# B4-B7: key params from Sheet2
sc(ws4, 4, 1, '关键参数（引用盈亏平衡计算）', font=b_font)
sc(ws4, 5, 1, 'BEP收入');          sc(ws4, 5, 2, f"='{SN_BEP}'!B{BEP_REV_ROW}", fmt=nfmt)
sc(ws4, 6, 1, '需弥补缺口');        sc(ws4, 6, 2, f"='{SN_BEP}'!B{BEP_GAP_ROW}", fmt=nfmt)
sc(ws4, 7, 1, '变动成本率(1-CMR)'); sc(ws4, 7, 2, f"=1-'{SN_BEP}'!B{BEP_CMR_ROW}", fmt=pfmt)
sc(ws4, 8, 1, '当前120&155收入');   sc(ws4, 8, 2, f"='{SN_PARAM}'!B{R_VAR_REV}", fmt=nfmt)

r = 10
hdr(ws4, r, ['产能利用率', '120&155收入', '变动成本', '需覆盖成本(VC+缺口)', '缺口线']); r += 1
A_DATA_START = r

for pct in [0, 20, 40, 60, 80, 100, 120, 140]:
    sc(ws4, r, 1, f'{pct}%')
    sc(ws4, r, 2, f'=B5*{pct/100}', fmt=nfmt)       # Revenue
    sc(ws4, r, 3, f'=B{r}*B7', fmt=nfmt)             # VC
    sc(ws4, r, 4, f'=C{r}+B6', fmt=nfmt)             # TC = VC + Gap
    sc(ws4, r, 5, f'=B6', fmt=nfmt)                  # Gap line
    r += 1
A_DATA_END = r - 1

# BEP marker row
sc(ws4, r, 1, 'v BEP点')
sc(ws4, r, 2, f'=B5', fmt=nfmt); sc(ws4, r, 3, f'=B5*B7', fmt=nfmt); sc(ws4, r, 4, f'=B5', fmt=nfmt); sc(ws4, r, 5, f'=B6', fmt=nfmt); r += 1

# Current position row
sc(ws4, r, 1, '* 当前位置')
sc(ws4, r, 2, f'=B8', fmt=nfmt); sc(ws4, r, 3, f'=B{r}*B7', fmt=nfmt); sc(ws4, r, 4, f'=C{r}+B6', fmt=nfmt); sc(ws4, r, 5, f'=B6', fmt=nfmt)
r += 2

# Chart A
chart_a = LineChart()
chart_a.title = 'A. 本量利图 - 120&155合计'
chart_a.y_axis.title = '金额 (元)'; chart_a.x_axis.title = '产能利用率 (100%=BEP)'
chart_a.style = 2; chart_a.width = 24; chart_a.height = 15

cats_a = Reference(ws4, min_col=1, min_row=A_DATA_START, max_row=A_DATA_END)
chart_a.add_data(Reference(ws4, min_col=2, min_row=A_DATA_START-1, max_row=A_DATA_END), titles_from_data=True)
chart_a.add_data(Reference(ws4, min_col=4, min_row=A_DATA_START-1, max_row=A_DATA_END), titles_from_data=True)
chart_a.add_data(Reference(ws4, min_col=5, min_row=A_DATA_START-1, max_row=A_DATA_END), titles_from_data=True)
chart_a.set_categories(cats_a)
chart_a.series[0].graphicalProperties.line.width = 28500
chart_a.series[1].graphicalProperties.line.width = 28500
chart_a.series[2].graphicalProperties.line.width = 15000
chart_a.series[2].graphicalProperties.line.dashStyle = 'dash'
chart_a.legend.position = 'b'
ca_row = r + 1
ws4.add_chart(chart_a, f'A{ca_row}')
r = ca_row + 19

# ==== Chart B: Profit Curve ====
ws4.cell(r, 1, '图B: 利润-收入曲线数据').font = s_font; r += 1
hdr(ws4, r, ['120&155收入', '利润', '零利润线', '标记']); r += 1
B_HDR_ROW = r - 1
B_DATA_START = r

for rev_m in range(0, 10500000, 500000):
    sc(ws4, r, 1, rev_m, fmt=nfmt)
    sc(ws4, r, 2, f"=A{r}*'{SN_BEP}'!B{BEP_CMR_ROW}-'{SN_BEP}'!B{BEP_GAP_ROW}", fmt=nfmt)
    sc(ws4, r, 3, 0, fmt=nfmt)
    sc(ws4, r, 4, '')
    r += 1
B_DATA_END = r - 1

# BEP point
sc(ws4, r, 1, f"='{SN_BEP}'!B{BEP_REV_ROW}", fmt=nfmt)
sc(ws4, r, 2, 0, fmt=nfmt); sc(ws4, r, 3, 0, fmt=nfmt); sc(ws4, r, 4, 'BEP')
B_BEP_ROW = r; r += 1
# Current point
sc(ws4, r, 1, f"='{SN_PARAM}'!B{R_VAR_REV}", fmt=nfmt)
sc(ws4, r, 2, f"=A{r}*'{SN_BEP}'!B{BEP_CMR_ROW}-'{SN_BEP}'!B{BEP_GAP_ROW}", fmt=nfmt)
sc(ws4, r, 3, 0, fmt=nfmt); sc(ws4, r, 4, '当前')
B_CUR_ROW = r; r += 2

# Chart B
chart_b = LineChart()
chart_b.title = 'B. 利润-收入曲线 - 120&155合计'
chart_b.y_axis.title = '利润 (元)'; chart_b.x_axis.title = '120&155收入 (元)'
chart_b.style = 2; chart_b.width = 24; chart_b.height = 15

cats_b = Reference(ws4, min_col=1, min_row=B_DATA_START, max_row=B_DATA_END)
chart_b.add_data(Reference(ws4, min_col=2, min_row=B_HDR_ROW, max_row=B_DATA_END), titles_from_data=True)
chart_b.add_data(Reference(ws4, min_col=3, min_row=B_HDR_ROW, max_row=B_DATA_END), titles_from_data=True)
chart_b.set_categories(cats_b)
chart_b.series[0].graphicalProperties.line.width = 28500
chart_b.series[1].graphicalProperties.line.width = 15000
chart_b.series[1].graphicalProperties.line.dashStyle = 'dash'

# BEP marker
chart_b.add_data(Reference(ws4, min_col=2, min_row=B_BEP_ROW, max_row=B_BEP_ROW))
chart_b.series[2].graphicalProperties.line.noFill = True
chart_b.series[2].marker.symbol = 'diamond'; chart_b.series[2].marker.size = 10
chart_b.series[2].marker.graphicalProperties.solidFill = 'FF0000'
chart_b.series[2].title = SeriesLabel(v='BEP')

# Current marker
chart_b.add_data(Reference(ws4, min_col=2, min_row=B_CUR_ROW, max_row=B_CUR_ROW))
chart_b.series[3].graphicalProperties.line.noFill = True
chart_b.series[3].marker.symbol = 'circle'; chart_b.series[3].marker.size = 10
chart_b.series[3].marker.graphicalProperties.solidFill = '2F5496'
chart_b.series[3].title = SeriesLabel(v='当前')

chart_b.legend.position = 'b'
cb_row = r + 1
ws4.add_chart(chart_b, f'A{cb_row}')
r = cb_row + 19

# ==== Chart C: Iso-profit line ====
ws4.cell(r, 1, '图C: 等利润线数据').font = s_font; r += 1

sc(ws4, r, 1, '关键参数', font=b_font)
sc(ws4, r, 2, f"='{SN_BEP}'!B{BEP_GAP_ROW}", font=b_font, fmt=nfmt); r += 1
sc(ws4, r, 1, '* 需弥补缺口')
sc(ws4, r, 2, f"='{SN_BEP}'!B{BEP_GAP_ROW}", fmt=nfmt); r += 1
sc(ws4, r, 1, '120单位毛利')
sc(ws4, r, 2, f"='{SN_PARAM}'!B{R_UMARGIN}", fmt=nfmt); r += 1
sc(ws4, r, 1, '155单位毛利')
sc(ws4, r, 2, f"='{SN_PARAM}'!C{R_UMARGIN}", fmt=nfmt); r += 2

hdr(ws4, r, ['155年销量', '120年销量(BEP线)', '利润(验证~0)']); r += 1
C_HDR_ROW = r - 1
C_DATA_START = r

for q155 in range(0, 1401, 50):
    sc(ws4, r, 1, q155, fmt=nfmt)
    sc(ws4, r, 2,
        f"=MAX(0,ROUND(("
        f"'{SN_BEP}'!B{BEP_GAP_ROW}-A{r}*'{SN_PARAM}'!C{R_UMARGIN}"
        f")/'{SN_PARAM}'!B{R_UMARGIN},0))",
        fmt=nfmt)
    sc(ws4, r, 3,
        f"=B{r}*'{SN_PARAM}'!B{R_UMARGIN}+A{r}*'{SN_PARAM}'!C{R_UMARGIN}-'{SN_BEP}'!B{BEP_GAP_ROW}",
        fmt=nfmt)
    r += 1
C_DATA_END = r - 1

# Current
sc(ws4, r, 1, f"='{SN_PARAM}'!C{R_SALES}", fmt=nfmt)
sc(ws4, r, 2, f"='{SN_PARAM}'!B{R_SALES}", fmt=nfmt)
sc(ws4, r, 3, f"=B{r}*'{SN_PARAM}'!B{R_UMARGIN}+A{r}*'{SN_PARAM}'!C{R_UMARGIN}-'{SN_BEP}'!B{BEP_GAP_ROW}", fmt=nfmt)
C_CUR_ROW = r; r += 1

# Endpoint: Pure 155
sc(ws4, r, 1, f"=ROUNDUP('{SN_BEP}'!B{BEP_REV_ROW}/'{SN_PARAM}'!C{R_PRICE},0)", fmt=nfmt)
sc(ws4, r, 2, 0, fmt=nfmt)
sc(ws4, r, 3, f"=B{r}*'{SN_PARAM}'!B{R_UMARGIN}+A{r}*'{SN_PARAM}'!C{R_UMARGIN}-'{SN_BEP}'!B{BEP_GAP_ROW}", fmt=nfmt)
C_EP155_ROW = r; r += 1

# Endpoint: Pure 120
sc(ws4, r, 1, 0, fmt=nfmt)
sc(ws4, r, 2, f"=ROUNDUP('{SN_BEP}'!B{BEP_REV_ROW}/'{SN_PARAM}'!B{R_PRICE},0)", fmt=nfmt)
sc(ws4, r, 3, f"=B{r}*'{SN_PARAM}'!B{R_UMARGIN}+A{r}*'{SN_PARAM}'!C{R_UMARGIN}-'{SN_BEP}'!B{BEP_GAP_ROW}", fmt=nfmt)
C_EP120_ROW = r; r += 2

# Chart C: Scatter
chart_c = ScatterChart()
chart_c.title = 'C. 双产品BEP等利润线（线上=保本  线上方=盈利）'
chart_c.y_axis.title = '120年销量 (个)'; chart_c.x_axis.title = '155年销量 (个)'
chart_c.style = 2; chart_c.width = 24; chart_c.height = 15

x_c = Reference(ws4, min_col=1, min_row=C_DATA_START, max_row=C_DATA_END)
y_c = Reference(ws4, min_col=2, min_row=C_DATA_START, max_row=C_DATA_END)
chart_c.add_data(y_c, titles_from_data=False)
chart_c.set_categories(x_c)
chart_c.series[0].title = SeriesLabel(v='BEP线 (利润=0)')
chart_c.series[0].graphicalProperties.line.width = 25000
chart_c.series[0].graphicalProperties.line.solidFill = '2F5496'

# Current point
chart_c.add_data(Reference(ws4, min_col=2, min_row=C_CUR_ROW, max_row=C_CUR_ROW))
chart_c.series[1].graphicalProperties.line.noFill = True
chart_c.series[1].marker.symbol = 'circle'; chart_c.series[1].marker.size = 14
chart_c.series[1].marker.graphicalProperties.solidFill = 'FF0000'
chart_c.series[1].title = SeriesLabel(v='当前位置')

# EP120
chart_c.add_data(Reference(ws4, min_col=2, min_row=C_EP120_ROW, max_row=C_EP120_ROW))
chart_c.series[2].graphicalProperties.line.noFill = True
chart_c.series[2].marker.symbol = 'diamond'; chart_c.series[2].marker.size = 9
chart_c.series[2].marker.graphicalProperties.solidFill = '2F5496'
chart_c.series[2].title = SeriesLabel(v='纯120端点')

# EP155
chart_c.add_data(Reference(ws4, min_col=2, min_row=C_EP155_ROW, max_row=C_EP155_ROW))
chart_c.series[3].graphicalProperties.line.noFill = True
chart_c.series[3].marker.symbol = 'diamond'; chart_c.series[3].marker.size = 9
chart_c.series[3].marker.graphicalProperties.solidFill = '00B050'
chart_c.series[3].title = SeriesLabel(v='纯155端点')

chart_c.legend.position = 'b'
cc_row = r + 1
ws4.add_chart(chart_c, f'A{cc_row}')

# ================================================================
# Save
# ================================================================
out_path = r'2025年年报\盈亏平衡测算模型.xlsx'
wb.save(out_path)

# Console summary
um120 = 2600 - 1170
um155 = 4000 - 1800
fc_total = sum(m for _, m in fc_items) * 12
trade_gm = 600000
gap = fc_total - trade_gm
cmr = (500*um120 + 500*um155) / (500*2600 + 500*4000)
bep_rev = gap / cmr

print(f'[OK] Saved: {out_path}')
print('='*55)
print(f'  Fixed costs/yr: {fc_total:,}')
print(f'  Trade GM: {trade_gm:,}')
print(f'  Gap needed: {gap:,}')
print(f'  120&155 CMR: {cmr*100:.1f}%')
print(f'  BEP Revenue: {bep_rev:,.0f}')
print(f'  Pure 120: {int(bep_rev/2600)+1} units  Pure 155: {int(bep_rev/4000)+1} units')
print(f'  Default (500 each): Total GM={500*um120+500*um155+trade_gm:,}  Profit={500*um120+500*um155+trade_gm-fc_total:,}')
